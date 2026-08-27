
"""Service for importing player and match data from Excel workbooks and OpenGotha XML files."""
from pathlib import Path
import logging

import openpyxl

from config import DEFAULT_RATING, DEFAULT_RD, DEFAULT_VOLATILITY
from services.common import get_db
from services.helpers import (
    header_index,
    looks_like_player_name,
    normalize_key,
    normalize_round_note,
    normalize_round_note_for_storage,
    normalize_text,
    parse_date_value,
    should_skip_sheet,
)
from services.player_service import build_player_lookup, ensure_player
from services.import_gotha import GothaMatch, parse_gotha_xml as _parse_gotha_xml
from services.rating_service import clear_dirty_date, mark_dirty
from services.tournament_service import _suggest_player_name, read_gotha_tournament

logger = logging.getLogger(__name__)


def parse_gotha_xml(xml_path) -> list[GothaMatch]:
    """Explicit public re-export for the shared OpenGotha parser."""
    return _parse_gotha_xml(xml_path)


def build_import_preview(conn, xml_path):
    """Return an actionable OpenGotha reconciliation preview before commit."""
    metadata = read_gotha_tournament(xml_path)
    player_rows = conn.execute(
        "SELECT id, display_name, first_name, last_name FROM players WHERE active = 1"
    ).fetchall()
    lookup = {
        normalize_key(row["display_name"]): row["id"]
        for row in player_rows
        if row["display_name"]
    }

    rows = []
    summary = {
        "players_total": 0,
        "exact_matches": 0,
        "fuzzy_matches": 0,
        "new_players": 0,
        "duplicate_players": 0,
        "unresolved_players": 0,
        "metadata_mismatches": 0,
    }
    seen_keys = set()

    for participant in metadata["players"]:
        display_name = participant.get("display_name") or ""
        source_key = participant.get("key") or normalize_key(display_name)
        resolved_player_id = lookup.get(normalize_key(display_name))
        status = "new"
        suggested_name = None
        suggested_player_id = None

        if source_key in seen_keys:
            status = "duplicate"
            summary["duplicate_players"] += 1
        elif resolved_player_id is not None:
            status = "exact"
            summary["exact_matches"] += 1
        else:
            suggested_name = _suggest_player_name(display_name, conn)
            if suggested_name:
                suggestion_match = conn.execute(
                    "SELECT id, display_name FROM players WHERE active = 1 AND display_name = ?",
                    (suggested_name,),
                ).fetchone()
                if suggestion_match is not None:
                    suggested_player_id = suggestion_match["id"]
                    status = "fuzzy"
                    summary["fuzzy_matches"] += 1
                else:
                    status = "new"
            summary["new_players"] += 1 if status == "new" else 0
            summary["unresolved_players"] += 1 if status == "new" else 0

        seen_keys.add(source_key)

        rows.append(
            {
                "display_name": display_name,
                "status": status,
                "resolved_player_id": resolved_player_id,
                "suggested_name": suggested_name,
                "suggested_player_id": suggested_player_id,
                "source_key": source_key,
                "decision": "reject" if status == "duplicate" else (
                    str(resolved_player_id or suggested_player_id or "new")
                ),
            }
        )

    summary["players_total"] = len(rows)
    metadata_mismatches = []
    if _table_exists(conn, "tournaments"):
        existing_tournament = conn.execute(
            """
            SELECT id, name, begin_date
            FROM tournaments
            WHERE lower(name) = lower(?) AND begin_date = ?
            ORDER BY id DESC LIMIT 1
            """,
            (metadata.get("name", ""), metadata.get("begin_date", "")),
        ).fetchone()
        if existing_tournament is not None:
            metadata_mismatches.append({
                "field": "tournament",
                "imported": metadata.get("name"),
                "existing": existing_tournament["name"],
                "reason": "A tournament with the same name and start date already exists.",
            })
            summary["metadata_mismatches"] = 1
    return {
        "metadata": {
            "name": metadata.get("name"),
            "begin_date": metadata.get("begin_date"),
            "rounds": metadata.get("rounds"),
            "pairing_system": metadata.get("pairing_system"),
        },
        "metadata_mismatches": metadata_mismatches,
        "summary": summary,
        "rows": rows,
    }


def _table_exists(conn, table_name):
    return conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone() is not None


def reset_import_tables(conn):
    conn.execute("DELETE FROM rating_snapshots")
    conn.execute("DELETE FROM matches")
    conn.execute("DELETE FROM players")
    conn.commit()


def import_workbook_data(file_path, reset=False):
    path = Path(file_path)

    if not path.exists():
        return {"players": 0, "matches": 0}

    workbook = openpyxl.load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    conn = get_db()

    if reset:
        reset_import_tables(conn)
        clear_dirty_date(conn=conn)

    player_lookup = build_player_lookup(conn)
    players_imported = 0
    matches_imported = 0
    earliest_match_date = None

    current_ratings = {}
    ratings_sheet_name = "Ratings by Player"

    if ratings_sheet_name in workbook.sheetnames:
        ratings_rows = list(workbook[ratings_sheet_name].iter_rows(values_only=True))
        if ratings_rows:
            ratings_headers = [normalize_text(cell) for cell in ratings_rows[0]]
            ratings_name_idx = header_index(ratings_headers, ["name", "player", "player name", "jugador"])
            ratings_rating_idx = header_index(ratings_headers, ["glicko", "rating"])

            for row in ratings_rows[1:]:
                try:
                    if not row:
                        continue
                    player_name = (
                        row[ratings_name_idx]
                        if ratings_name_idx is not None and ratings_name_idx < len(row)
                        else None
                    )
                    if player_name is None or not looks_like_player_name(player_name):
                        continue
                    if (
                        ratings_rating_idx is not None
                        and ratings_rating_idx < len(row)
                        and row[ratings_rating_idx] not in (None, "")
                    ):
                        current_ratings[str(player_name).strip()] = float(row[ratings_rating_idx])
                except Exception:
                    logger.warning("Skipping invalid rating row: %r", row, exc_info=True)
                    continue

    player_sheet_name = "Player List"

    if player_sheet_name in workbook.sheetnames:
        rows = list(workbook[player_sheet_name].iter_rows(values_only=True))
        if rows:
            headers = [normalize_text(cell) for cell in rows[0]]
            name_idx = header_index(headers, ["player name", "name", "player", "jugador"])
            rating_idx = header_index(headers, ["player's rating", "players rating", "initial rating"])

            for row in rows[1:]:
                try:
                    if not row:
                        continue

                    display_name = (
                        row[name_idx]
                        if name_idx is not None and name_idx < len(row)
                        else None
                    )
                    if display_name is None or not looks_like_player_name(display_name):
                        continue

                    initial_rating = DEFAULT_RATING
                    if rating_idx is not None and rating_idx < len(row):
                        try:
                            if row[rating_idx] not in (None, ""):
                                initial_rating = float(row[rating_idx])
                        except Exception:
                            logger.warning(
                                "Invalid initial rating %r; using the default",
                                row[rating_idx],
                                exc_info=True,
                            )

                    current_rating = current_ratings.get(str(display_name).strip(), initial_rating)
                    ensure_player(conn, display_name, current_rating, initial_rating=initial_rating, player_lookup=player_lookup)
                    players_imported += 1
                except Exception:
                    logger.warning("Skipping player import row: %r", row, exc_info=True)
                    continue

    match_sheet_name = "Matches"
    if match_sheet_name in workbook.sheetnames:
        rows = list(workbook[match_sheet_name].iter_rows(values_only=True))
        if rows:
            headers = [normalize_text(cell) for cell in rows[0]]
            date_idx = header_index(headers, ["date", "fecha"])
            time_idx = header_index(headers, ["time", "hora", "round"])
            white_idx = header_index(headers, ["white", "blanco"])
            black_idx = header_index(headers, ["black", "negro"])
            result_idx = header_index(headers, ["winner", "ganador", "result", "resultado"])
            comment_idx = header_index(headers, ["comments", "comentarios", "event", "evento", "notes"])

            for row in rows[1:]:
                try:
                    if not row:
                        continue

                    white_name = row[white_idx] if white_idx is not None and white_idx < len(row) else None
                    black_name = row[black_idx] if black_idx is not None and black_idx < len(row) else None
                    winner = row[result_idx] if result_idx is not None and result_idx < len(row) else None

                    if not white_name or not black_name:
                        continue
                    if not looks_like_player_name(white_name) or not looks_like_player_name(black_name):
                        continue

                    white_id = ensure_player(conn, white_name, player_lookup=player_lookup)
                    black_id = ensure_player(conn, black_name, player_lookup=player_lookup)

                    date_value = row[date_idx] if date_idx is not None and date_idx < len(row) else None
                    match_date = parse_date_value(date_value)

                    time_value = row[time_idx] if time_idx is not None and time_idx < len(row) else None
                    event = row[comment_idx] if comment_idx is not None and comment_idx < len(row) else None

                    winner_text = normalize_text(winner).lower()
                    white_key = normalize_key(white_name)
                    black_key = normalize_key(black_name)
                    winner_key = normalize_key(winner)

                    if "draw" in winner_text or "empate" in winner_text or "½" in winner_text or "1/2" in winner_text:
                        result = "1/2-1/2"
                    elif winner_key == white_key:
                        result = "1-0"
                    elif winner_key == black_key:
                        result = "0-1"
                    else:
                        result = "1/2-1/2"

                    existing = conn.execute(
                        """
                        SELECT id
                        FROM matches
                        WHERE match_date = ?
                        AND white_player_id = ?
                        AND black_player_id = ?
                        AND result = ?
                        """,
                        (match_date, white_id, black_id, result),
                    ).fetchone()

                    if existing is None:
                        raw_note = normalize_round_note_for_storage(time_value)
                        conn.execute(
                            """
                            INSERT INTO matches
                            (
                                match_date,
                                white_player_id,
                                black_player_id,
                                result,
                                event,
                                notes,
                                round_number
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                match_date,
                                white_id,
                                black_id,
                                result,
                                normalize_text(event),
                                raw_note,
                                normalize_round_note(time_value),
                            ),
                        )
                        matches_imported += 1
                        if earliest_match_date is None or match_date < earliest_match_date:
                            earliest_match_date = match_date
                except Exception:
                    logger.warning("Skipping match import row: %r", row, exc_info=True)
                    continue
    else:
        logger.warning("Matches sheet not found in workbook: %s", path)

    for sheetname in workbook.sheetnames:
        if should_skip_sheet(sheetname):
            continue

        ws = workbook[sheetname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx = None
        for idx, row in enumerate(rows):
            if any("player's rating" in str(cell).lower() for cell in row if cell is not None):
                header_row_idx = idx
                break

        if header_row_idx is None:
            continue

        headers = [normalize_text(cell) for cell in rows[header_row_idx]]
        date_idx = header_index(headers, ["date", "fecha"])
        rating_idx = header_index(headers, ["player's rating", "players rating"])

        if date_idx is None or rating_idx is None:
            continue

        player_row = conn.execute(
            """
            SELECT id, display_name, rating
            FROM players
            WHERE display_name = ?
            """,
            (sheetname,),
        ).fetchone()

        if player_row is None:
            logger.warning("Player not found while processing snapshot sheet %r", sheetname)
            continue

        player_id = player_row["id"]
        snapshot_rows = []

        for row in rows[header_row_idx + 1:]:
            if not row or all(cell is None for cell in row):
                break
            snapshot_rows.append(row)

        snapshot_rows.reverse()

        for row in snapshot_rows:
            if not row or all(cell is None for cell in row):
                break

            date_value = row[date_idx] if date_idx < len(row) else None
            rating_value = row[rating_idx] if rating_idx < len(row) else None
            if str(date_value).strip().lower() == "<initial rating>":
                continue

            try:
                snapshot_date = parse_date_value(date_value)
                rating = float(rating_value) if rating_value not in (None, "") else DEFAULT_RATING
            except ValueError:
                continue

            conn.execute(
                """
                INSERT INTO rating_snapshots (player_id, snapshot_date, rating, rd, volatility)
                VALUES (?, ?, ?, ?, ?)
                """,
                (player_id, snapshot_date, rating, DEFAULT_RD, DEFAULT_VOLATILITY),
            )

    conn.commit()
    conn.close()

    if earliest_match_date:
        mark_dirty(earliest_match_date)

    return {
        "players": players_imported,
        "matches": matches_imported,
    }


def import_gotha_xml(xml_path):

    conn = get_db()

    matches = parse_gotha_xml(xml_path)

    matches_imported = 0
    earliest_match_date = None

    for match in matches:

        white_id = ensure_player(
            conn,
            match["white"]
        )

        black_id = ensure_player(
            conn,
            match["black"]
        )

        existing = conn.execute(
            """
            SELECT id
            FROM matches
            WHERE match_date = ?
            AND white_player_id = ?
            AND black_player_id = ?
            AND result = ?
            """,
            (
                match["match_date"],
                white_id,
                black_id,
                match["result"],
            ),
        ).fetchone()

        if existing:
            continue

        conn.execute(
            """
            INSERT INTO matches
            (
                match_date,
                white_player_id,
                black_player_id,
                result,
                event,
                notes,
                round_number
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                match["match_date"],
                white_id,
                black_id,
                match["result"],
                match["event"],
                normalize_round_note_for_storage(
                    f"Round {match['round']}" if match.get("round") else ""
                ),
                normalize_round_note(match.get("round")),
            ),
        )

        matches_imported += 1
        if earliest_match_date is None or match["match_date"] < earliest_match_date:
            earliest_match_date = match["match_date"]

    conn.commit()
    conn.close()

    if earliest_match_date:
        mark_dirty(earliest_match_date)

    return {
        "players": 0,
        "matches": matches_imported,
    }
